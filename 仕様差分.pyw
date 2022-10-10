# モジュールのインポート
import os
import tkinter
import tkinter.filedialog
import tkinter.messagebox
from tkinter import filedialog
from tkinter import messagebox
#from tkinter import ttk
from openpyxl import load_workbook
#import sys
import traceback
import logging
logger = logging.getLogger()

def ask_folder():
    """ 参照ボタン"""
    fTyp = [("", "*")]
    iDir = os.path.abspath(os.path.dirname(__file__))
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")], defaultextension=".xlsx")
    folder_path.set(path)

def app():
    file_path = folder_path.get()
    file_path2 = folder_path2.get()
    sheet_path2 = sheet_path.get()
    print(file_path)
    print(file_path2)
    print(sheet_path2)
    #"C:\Users\Koichiro Ishida\Desktop\参考資料\2_B-Atlas_B-Titan_通常版_1st_テスト概略設計書_第10版.xlsx"
    try:
        wb = load_workbook(filename=file_path, data_only=True)

        cellrow = file_path2

        #ws = wb['テスト範囲分析マトリクス_本体機能']
        ws = wb[str(sheet_path2)]
        for row in ws.iter_rows(min_row=cellrow,max_row=cellrow, values_only=True):
            SearchA = [row[104]]
            print(SearchA)
        ws1 = wb['仕様差分リスト']

        #仕様差分の内容を検索してる
        b = 0
        #あかんかった場合をカウントして検索した行数の総数と等しい場合はエラー表示
        g = 0
        #仕様差分シートの検索
        for row in ws1["Q1:Q511"]:
            values = []
            
            for col in row:
                b +=1
                values.append(col.value)
                if SearchA == values: 
                    print(str(b) + "行目")
                    for row in ws1["E"+str(b) +":"+ "E"+str(b) ]:
                        valuesC = []
                        for col in row:
                            valuesC.append(col.value)
                            print("仕様差分：" + str(valuesC))
                            a = "仕様差分：" + str(valuesC)
                    messagebox.showinfo("完了", str(b) + "行目"+ "\r" + a )
                else :
                    #print("NO FOUND OUT")
                    g += 1
                
                if g == 511:
                    messagebox.showinfo("完了","仕様差分はありませんでした。" )           

        
    #エラー表示の表記
    except  Exception as e:
        traceback.print_exc()
        messagebox.showinfo("エラー", str(e))


#以下、Tkinterの設定
root = tkinter.Tk()
root.title(u"仕様差分サーチ")
root.geometry("625x210")

#ラベル
Static1 = tkinter.Label(text=u'差分リストを抽出します。')

folder_path = tkinter.StringVar()
folder_path2 = tkinter.IntVar()
sheet_path = tkinter.StringVar()

folder_label = tkinter.Label(text= "1.概略設計書の指定")

folder_box = tkinter.Entry(textvariable=folder_path,width = 100)


folder_btn = tkinter.Button(text="参照", command=ask_folder)

sheet_label = tkinter.Label(text= "2.マトリクスがあるシート名の指定")
sheet_box = tkinter.Entry(textvariable=sheet_path,width = 100)

Static2 = tkinter.Label(text=u'3.Yのある行数の指定')

folder_box2 = tkinter.Entry(textvariable=folder_path2,width = 8)

# ウィジェット（実行ボタン）
app_btn = tkinter.Button(text="実行", command=app)

Static1.pack(anchor = tkinter.NW, padx=10)
folder_label.pack(anchor = tkinter.NW, padx=10)
folder_box.pack(anchor = tkinter.N)
folder_btn.pack(anchor = tkinter.E, padx=10)
sheet_label.pack(anchor = tkinter.NW, padx=10)
sheet_box.pack(anchor = tkinter.N)
Static2.pack(anchor = tkinter.NW, padx=10)
folder_box2.pack(anchor = tkinter.NW, padx=10)
app_btn.pack(anchor = tkinter.E, padx=10)

root.mainloop()
